from __future__ import annotations

from copy import deepcopy
from pathlib import Path

from openpyxl import load_workbook

from sekretariat_app.sips.settings import MASTER_PERSONNEL_XLSX


class PersonnelMaster:
    """Master anggota DPRD dan ASN dari workbook resmi SIPS."""

    def __init__(self, path: str | Path = MASTER_PERSONNEL_XLSX) -> None:
        self.path = Path(path)
        self._dprd: list[dict[str, str]] | None = None
        self._asn: list[dict[str, str]] | None = None

    def _load(self) -> None:
        workbook = load_workbook(self.path, read_only=True, data_only=True)
        try:
            dprd_sheet = next(sheet for sheet in workbook.sheetnames if "dprd" in sheet.lower() or "anggota" in sheet.lower())
            asn_sheet = next(sheet for sheet in workbook.sheetnames if "asn" in sheet.lower() or "pendamping" in sheet.lower())
            self._dprd = []
            for row in workbook[dprd_sheet].iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
                name, position, category = (str(value).strip() if value is not None else "" for value in row)
                if name:
                    self._dprd.append({"nama": name, "jabatan": position, "kategori": category or "Lainnya"})
            self._asn = []
            for row in workbook[asn_sheet].iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
                name, nip, rank, position = (str(value).strip() if value is not None else "" for value in row)
                if name:
                    self._asn.append({"nama": name, "nip": nip, "pangkat": rank, "jabatan": position})
        finally:
            workbook.close()

    @property
    def dprd(self) -> list[dict[str, str]]:
        if self._dprd is None:
            self._load()
        return deepcopy(self._dprd or [])

    @property
    def asn(self) -> list[dict[str, str]]:
        if self._asn is None:
            self._load()
        return deepcopy(self._asn or [])

    @property
    def dprd_categories(self) -> list[str]:
        return list(dict.fromkeys(item["kategori"] for item in self.dprd if item["kategori"]))

    @property
    def dprd_signers(self) -> list[str]:
        rows = [item for item in self.dprd if item["kategori"].lower() == "pimpinan dprd"]
        return [f"{item['jabatan']} - {item['nama']}" for item in rows]

    @property
    def asn_signers(self) -> list[str]:
        rows = [item for item in self.asn if "sekretaris dprd" in item["jabatan"].lower()]
        rows = rows or self.asn
        return [f"{item['jabatan']} - {item['nama']}" for item in rows]
