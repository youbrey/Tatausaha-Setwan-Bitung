from __future__ import annotations

import ctypes
import os
import re
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PySide6.QtCore import QDate, QTimer, Qt, QUrl, Signal
from PySide6.QtGui import QDesktopServices
from PySide6.QtPrintSupport import QPrinterInfo
from PySide6.QtWidgets import (
    QApplication,
    QCheckBox,
    QComboBox,
    QCompleter,
    QDateEdit,
    QFileDialog,
    QFrame,
    QGridLayout,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QListWidget,
    QMessageBox,
    QPlainTextEdit,
    QPushButton,
    QScrollArea,
    QSplitter,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
)

from sekretariat_app.auth import User, UserRepository
from sekretariat_app.sips.constants import (
    DEFAULT_TRAVEL_DESTINATIONS,
    JENIS_PERJALANAN_DPRD,
    JENIS_PERJALANAN_SETWAN,
    JENIS_RAPAT_OPTIONS,
    PELAKSANA_RAPAT_CUSTOM,
    PELAKSANA_RAPAT_OPTIONS,
)
from sekretariat_app.sips.models import (
    DocumentRecord,
    InvitationFormData,
    TravelFormData,
    day_name_id,
    format_date_id,
)
from sekretariat_app.sips.repository import SIPSRepository
from sekretariat_app.sips.service import SIPSService
from sekretariat_app.ui.live_preview import LiveDocumentPreview


def _date_edit(value: date | None = None) -> QDateEdit:
    widget = QDateEdit()
    widget.setCalendarPopup(True)
    widget.setDisplayFormat("dd/MM/yyyy")
    current = value or date.today()
    widget.setDate(QDate(current.year, current.month, current.day))
    return widget


def _python_date(widget: QDateEdit) -> date:
    value = widget.date()
    return date(value.year(), value.month(), value.day())


def _set_date(widget: QDateEdit, value: date) -> None:
    widget.setDate(QDate(value.year, value.month, value.day))


def _lines(editor: QPlainTextEdit) -> list[str]:
    return [line.strip() for line in editor.toPlainText().splitlines() if line.strip()]


def _open_path(path: str | Path) -> None:
    QDesktopServices.openUrl(QUrl.fromLocalFile(str(Path(path).resolve())))


def _print_to(path: str | Path, printer_name: str) -> None:
    target = Path(path).resolve()
    if not target.exists():
        raise FileNotFoundError(target)
    if os.name != "nt":
        _open_path(target)
        return
    shell_execute = ctypes.windll.shell32.ShellExecuteW
    shell_execute.restype = ctypes.c_ssize_t
    result = shell_execute(
        None,
        "printto",
        str(target),
        f'"{printer_name}"' if printer_name else None,
        str(target.parent),
        0,
    )
    if result <= 32:
        raise OSError(f"Windows gagal mengirim dokumen ke printer (kode {result}).")


class PersonnelCheckList(QWidget):
    def __init__(self, rows: list[dict[str, str]], grouped: bool, title: str):
        super().__init__()
        self.rows = rows
        self.grouped = grouped
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        header = QLabel(title)
        header.setObjectName("SectionTitle")
        layout.addWidget(header)
        self.search = QLineEdit()
        self.search.setPlaceholderText("Cari nama, jabatan, atau kategori…")
        self.search.setClearButtonEnabled(True)
        self.search.textChanged.connect(self._filter)
        layout.addWidget(self.search)
        selection_actions = QHBoxLayout()
        select_visible = QPushButton("Centang Semua Tampil")
        clear_selection = QPushButton("Bersihkan")
        select_visible.clicked.connect(self.select_visible)
        clear_selection.clicked.connect(self.clear)
        selection_actions.addWidget(select_visible)
        selection_actions.addWidget(clear_selection)
        selection_actions.addStretch()
        layout.addLayout(selection_actions)
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(("Nama", "Jabatan/Kategori"))
        self.tree.header().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.tree.header().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.tree.setAlternatingRowColors(True)
        layout.addWidget(self.tree, 1)
        self._populate()

    def _populate(self) -> None:
        self.tree.clear()
        if self.grouped:
            groups: dict[str, QTreeWidgetItem] = {}
            for row in self.rows:
                category = row.get("kategori", "Lainnya") or "Lainnya"
                parent = groups.get(category)
                if parent is None:
                    parent = QTreeWidgetItem((category, ""))
                    parent.setFirstColumnSpanned(True)
                    parent.setFlags(parent.flags() | Qt.ItemFlag.ItemIsAutoTristate | Qt.ItemFlag.ItemIsUserCheckable)
                    parent.setCheckState(0, Qt.CheckState.Unchecked)
                    self.tree.addTopLevelItem(parent)
                    groups[category] = parent
                item = QTreeWidgetItem((row.get("nama", ""), row.get("jabatan", "")))
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                item.setCheckState(0, Qt.CheckState.Unchecked)
                item.setData(0, Qt.ItemDataRole.UserRole, row)
                parent.addChild(item)
            self.tree.expandAll()
        else:
            for row in self.rows:
                detail = row.get("jabatan", "")
                if row.get("nip"):
                    detail = f"{detail} · NIP {row['nip']}"
                item = QTreeWidgetItem((row.get("nama", ""), detail))
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                item.setCheckState(0, Qt.CheckState.Unchecked)
                item.setData(0, Qt.ItemDataRole.UserRole, row)
                self.tree.addTopLevelItem(item)

    def _iter_people(self):
        for index in range(self.tree.topLevelItemCount()):
            top = self.tree.topLevelItem(index)
            if self.grouped:
                for child_index in range(top.childCount()):
                    yield top.child(child_index)
            else:
                yield top

    def _filter(self, text: str) -> None:
        query = text.strip().casefold()
        for index in range(self.tree.topLevelItemCount()):
            top = self.tree.topLevelItem(index)
            if self.grouped:
                visible = False
                for child_index in range(top.childCount()):
                    child = top.child(child_index)
                    row = child.data(0, Qt.ItemDataRole.UserRole) or {}
                    match = not query or query in " ".join(str(value) for value in row.values()).casefold()
                    child.setHidden(not match)
                    visible = visible or match
                top.setHidden(not visible)
            else:
                row = top.data(0, Qt.ItemDataRole.UserRole) or {}
                top.setHidden(bool(query) and query not in " ".join(str(value) for value in row.values()).casefold())

    def selected(self) -> list[dict[str, str]]:
        return [
            dict(item.data(0, Qt.ItemDataRole.UserRole))
            for item in self._iter_people()
            if item.checkState(0) == Qt.CheckState.Checked
        ]

    def set_selected(self, rows: list[dict[str, Any]]) -> None:
        wanted = {
            (str(row.get("nama", "")), str(row.get("kategori", "")))
            for row in rows
        }
        for item in self._iter_people():
            row = item.data(0, Qt.ItemDataRole.UserRole) or {}
            key = (str(row.get("nama", "")), str(row.get("kategori", "")))
            item.setCheckState(0, Qt.CheckState.Checked if key in wanted else Qt.CheckState.Unchecked)

    def select_visible(self) -> None:
        for item in self._iter_people():
            if not item.isHidden():
                item.setCheckState(0, Qt.CheckState.Checked)

    def clear(self) -> None:
        for item in self._iter_people():
            item.setCheckState(0, Qt.CheckState.Unchecked)


class TravelPage(QWidget):
    changed = Signal()

    def __init__(
        self,
        mode: str,
        service: SIPSService,
        repository: SIPSRepository,
        user: User,
        audit: UserRepository,
    ) -> None:
        super().__init__()
        self.mode = mode
        self.service = service
        self.repository = repository
        self.user = user
        self.audit = audit
        self.record_id: str | None = None
        self._build_ui()
        self.reset_form()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 20)
        layout.setSpacing(14)
        title = QLabel("Perjalanan Dinas DPRD" if self.mode == "dprd" else "Perjalanan Dinas Sekretariat DPRD")
        title.setObjectName("PageTitle")
        subtitle = QLabel(
            "Buat Surat Tugas, Surat Pemberitahuan, SPD depan-belakang, dan Daftar Hadir dalam satu proses."
        )
        subtitle.setObjectName("Subtitle")
        layout.addWidget(title)
        layout.addWidget(subtitle)

        workspace = QSplitter(Qt.Orientation.Horizontal)
        splitter = QSplitter(Qt.Orientation.Horizontal)
        form_scroll = QScrollArea()
        form_scroll.setWidgetResizable(True)
        form_host = QWidget()
        form_layout = QVBoxLayout(form_host)
        form_layout.setContentsMargins(0, 0, 8, 0)
        form_layout.setSpacing(12)
        form_layout.addWidget(self._numbers_card())
        form_layout.addWidget(self._agenda_card())
        form_layout.addWidget(self._schedule_card())
        form_layout.addWidget(self._destination_card())
        form_layout.addWidget(self._signer_card())
        form_layout.addStretch()
        form_scroll.setWidget(form_host)
        splitter.addWidget(form_scroll)

        personnel_card = QFrame()
        personnel_card.setObjectName("Card")
        personnel_layout = QVBoxLayout(personnel_card)
        personnel_layout.setContentsMargins(16, 16, 16, 16)
        tabs = QTabWidget()
        if self.mode == "dprd":
            self.dprd_list = PersonnelCheckList(self.service.master.dprd, True, "Anggota DPRD")
            self.asn_list = PersonnelCheckList(self.service.master.asn, False, "Pendamping ASN")
            tabs.addTab(self.dprd_list, "Anggota DPRD")
            tabs.addTab(self.asn_list, "Pendamping ASN")
            self.executor_list = self.companion_list = None
        else:
            self.executor_list = PersonnelCheckList(self.service.master.asn, False, "Pelaksana ASN")
            self.companion_list = PersonnelCheckList(self.service.master.asn, False, "Pendamping ASN")
            tabs.addTab(self.executor_list, "Pelaksana ASN")
            tabs.addTab(self.companion_list, "Pendamping ASN")
            self.dprd_list = self.asn_list = None
        personnel_layout.addWidget(tabs)
        splitter.addWidget(personnel_card)
        splitter.setStretchFactor(0, 3)
        splitter.setStretchFactor(1, 2)
        splitter.setChildrenCollapsible(False)
        workspace.addWidget(splitter)

        self.live_preview = LiveDocumentPreview()
        workspace.addWidget(self.live_preview)
        workspace.setStretchFactor(0, 5)
        workspace.setStretchFactor(1, 3)
        workspace.setChildrenCollapsible(False)
        workspace.setSizes((760, 440))
        layout.addWidget(workspace, 1)

        actions = QHBoxLayout()
        self.state_label = QLabel("Formulir baru")
        self.state_label.setObjectName("StatusLabel")
        actions.addWidget(self.state_label, 1)
        reset = QPushButton("Formulir Baru")
        draft = QPushButton("Simpan Draft")
        generate = QPushButton("Buat Semua Dokumen")
        generate.setObjectName("PrimaryButton")
        reset.clicked.connect(self.reset_form)
        draft.clicked.connect(self.save_draft)
        generate.clicked.connect(self.generate)
        actions.addWidget(reset)
        actions.addWidget(draft)
        actions.addWidget(generate)
        layout.addLayout(actions)
        self._connect_live_preview_signals()

    @staticmethod
    def _card(title: str) -> tuple[QFrame, QGridLayout]:
        card = QFrame()
        card.setObjectName("Card")
        layout = QGridLayout(card)
        layout.setContentsMargins(18, 16, 18, 18)
        heading = QLabel(title)
        heading.setObjectName("SectionTitle")
        layout.addWidget(heading, 0, 0, 1, 2)
        return card, layout

    def _numbers_card(self) -> QFrame:
        card, layout = self._card("Nomor Dokumen")
        self.number_fields: dict[str, QLineEdit] = {}
        entries = (
            (
                ("surat_tugas_dprd", "Surat Tugas DPRD"),
                ("surat_tugas_asn", "Surat Tugas Setwan"),
                ("pemberitahuan_dprd", "Pemberitahuan DPRD"),
                ("pemberitahuan_asn", "Pemberitahuan Setwan"),
                ("spd_dprd", "SPD DPRD"),
                ("spd_asn", "SPD Setwan/ASN"),
            )
            if self.mode == "dprd" else
            (
                ("surat_tugas_asn", "Surat Tugas Setwan"),
                ("pemberitahuan_asn", "Surat Pemberitahuan Setwan"),
                ("spd_pelaksana", "SPD Pelaksana ASN"),
                ("spd_pendamping", "SPD Pendamping ASN"),
            )
        )
        for index, (key, label) in enumerate(entries):
            row, column = divmod(index, 2)
            field = QLineEdit()
            field.setPlaceholderText(f"Masukkan nomor {label.lower()}…")
            layout.addWidget(QLabel(label), row * 2 + 1, column)
            layout.addWidget(field, row * 2 + 2, column)
            self.number_fields[key] = field
        return card

    def _agenda_card(self) -> QFrame:
        card, layout = self._card("Dasar dan Materi Kegiatan")
        self.travel_type = QComboBox()
        self.travel_type.addItems(JENIS_PERJALANAN_DPRD if self.mode == "dprd" else JENIS_PERJALANAN_SETWAN)
        self.basis_dprd = QPlainTextEdit()
        self.basis_dprd.setMaximumHeight(70)
        self.basis_asn = QPlainTextEdit()
        self.basis_asn.setMaximumHeight(70)
        self.subject = QPlainTextEdit()
        self.subject.setMaximumHeight(84)
        self.notice_subject = QPlainTextEdit()
        self.notice_subject.setMaximumHeight(84)
        widgets = (
            ("Jenis perjalanan", self.travel_type),
            ("Dasar Surat Tugas DPRD", self.basis_dprd),
            ("Dasar Surat Tugas ASN", self.basis_asn),
            ("Materi Surat Tugas & SPD", self.subject),
            ("Isi Surat Pemberitahuan", self.notice_subject),
        )
        for index, (label, widget) in enumerate(widgets, start=1):
            layout.addWidget(QLabel(label), index, 0)
            layout.addWidget(widget, index, 1)
        self.duplicate_title_warning = QLabel()
        self.duplicate_title_warning.setStyleSheet("color: #B45309; font-weight: 600;")
        self.duplicate_title_warning.setWordWrap(True)
        self.duplicate_title_warning.hide()
        layout.addWidget(self.duplicate_title_warning, len(widgets) + 1, 0, 1, 2)
        self._duplicate_title_timer = QTimer(self)
        self._duplicate_title_timer.setSingleShot(True)
        self._duplicate_title_timer.setInterval(700)
        self._duplicate_title_timer.timeout.connect(self._check_duplicate_title)
        self.subject.textChanged.connect(self._duplicate_title_timer.start)
        return card

    def _schedule_card(self) -> QFrame:
        card, layout = self._card("Tanggal Pelaksanaan")
        self.letter_date = _date_edit()
        self.start_date = _date_edit()
        self.end_date = _date_edit()
        self.duration = QLabel("1 hari")
        for widget in (self.start_date, self.end_date):
            widget.dateChanged.connect(self._update_duration)
        for row, (label, widget) in enumerate(
            (("Tanggal surat", self.letter_date), ("Tanggal mulai", self.start_date), ("Tanggal selesai", self.end_date), ("Lama perjalanan", self.duration)),
            start=1,
        ):
            layout.addWidget(QLabel(label), row, 0)
            layout.addWidget(widget, row, 1)
        return card

    def _destination_card(self) -> QFrame:
        card, layout = self._card("Tujuan Perjalanan")
        self.destination_input = QLineEdit()
        self.destination_input.setPlaceholderText("Contoh: DPRD Kota Manado atau Kota Bandung")
        completer = QCompleter(DEFAULT_TRAVEL_DESTINATIONS, self.destination_input)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        completer.setFilterMode(Qt.MatchFlag.MatchContains)
        completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        self.destination_input.setCompleter(completer)
        add = QPushButton("Tambah")
        add.clicked.connect(self._add_destination)
        self.destination_input.returnPressed.connect(self._add_destination)
        self.destinations = QListWidget()
        self.destinations.setMaximumHeight(112)
        remove = QPushButton("Hapus Tujuan Terpilih")
        remove.clicked.connect(lambda: self.destinations.takeItem(self.destinations.currentRow()))
        layout.addWidget(self.destination_input, 1, 0)
        layout.addWidget(add, 1, 1)
        layout.addWidget(self.destinations, 2, 0, 1, 2)
        layout.addWidget(remove, 3, 1)
        return card

    def _signer_card(self) -> QFrame:
        card, layout = self._card("Penandatangan")
        self.signer_dprd = QComboBox()
        self.signer_dprd.setEditable(True)
        self.signer_dprd.addItems(self.service.master.dprd_signers)
        self.signer_asn = QComboBox()
        self.signer_asn.setEditable(True)
        self.signer_asn.addItems(self.service.master.asn_signers)
        layout.addWidget(QLabel("Penandatangan DPRD"), 1, 0)
        layout.addWidget(self.signer_dprd, 1, 1)
        layout.addWidget(QLabel("Penandatangan ASN/SPD"), 2, 0)
        layout.addWidget(self.signer_asn, 2, 1)
        if self.mode == "setwan":
            self.signer_dprd.setVisible(False)
            layout.itemAtPosition(1, 0).widget().setVisible(False)
        return card

    def _connect_live_preview_signals(self) -> None:
        for field in self.number_fields.values():
            field.textChanged.connect(self._schedule_live_preview)
        for field in (
            self.basis_dprd,
            self.basis_asn,
            self.subject,
            self.notice_subject,
        ):
            field.textChanged.connect(self._schedule_live_preview)
        for combo in (self.travel_type, self.signer_dprd, self.signer_asn):
            combo.currentTextChanged.connect(self._schedule_live_preview)
        for calendar in (self.letter_date, self.start_date, self.end_date):
            calendar.dateChanged.connect(self._schedule_live_preview)
        destination_model = self.destinations.model()
        destination_model.rowsInserted.connect(self._schedule_live_preview)
        destination_model.rowsRemoved.connect(self._schedule_live_preview)
        destination_model.dataChanged.connect(self._schedule_live_preview)
        for listing in (
            self.dprd_list,
            self.asn_list,
            self.executor_list,
            self.companion_list,
        ):
            if listing:
                listing.tree.itemChanged.connect(self._schedule_live_preview)

    def _schedule_live_preview(self, *_args) -> None:
        data = self.collect()
        try:
            data.validate_preview()
        except ValueError as exc:
            self.live_preview.show_waiting(f"Live preview menunggu: {exc}")
            return
        self.live_preview.schedule(
            lambda output, current=data: self.service.generate_travel(current, output, preview=True)
        )

    def _check_duplicate_title(self) -> None:
        duplicate = self.repository.find_duplicate_travel_title(
            self.subject.toPlainText(), self.record_id,
        )
        if duplicate is None:
            self.duplicate_title_warning.clear()
            self.duplicate_title_warning.hide()
            return
        self.duplicate_title_warning.setText(
            "Materi ini sudah pernah dibuat pada "
            f"{duplicate.document_date or '-'} oleh {duplicate.author}. Periksa sebelum melanjutkan."
        )
        self.duplicate_title_warning.show()

    def _add_destination(self) -> None:
        value = self.destination_input.text().strip()
        if value:
            self.destinations.addItem(value)
            self.destination_input.clear()

    def _update_duration(self) -> None:
        days = max(1, (_python_date(self.end_date) - _python_date(self.start_date)).days + 1)
        self.duration.setText(f"{days} hari")

    def collect(self) -> TravelFormData:
        destinations = [self.destinations.item(index).text() for index in range(self.destinations.count())]
        return TravelFormData(
            mode=self.mode,
            document_numbers={key: field.text().strip() for key, field in self.number_fields.items()},
            letter_date=_python_date(self.letter_date),
            start_date=_python_date(self.start_date),
            end_date=_python_date(self.end_date),
            travel_type=self.travel_type.currentText(),
            basis_dprd=self.basis_dprd.toPlainText(),
            basis_asn=self.basis_asn.toPlainText(),
            subject=self.subject.toPlainText(),
            notice_subject=self.notice_subject.toPlainText(),
            destinations=destinations,
            signer_dprd=self.signer_dprd.currentText(),
            signer_asn=self.signer_asn.currentText(),
            dprd=self.dprd_list.selected() if self.dprd_list else [],
            asn=self.asn_list.selected() if self.asn_list else [],
            executors=self.executor_list.selected() if self.executor_list else [],
            companions=self.companion_list.selected() if self.companion_list else [],
        )

    def _save(self, data: TravelFormData, status: str, files: list[Path] = None) -> None:
        primary = data.document_numbers.get("surat_tugas_dprd" if self.mode == "dprd" else "surat_tugas_asn", "")
        self.record_id = self.repository.save(
            record_type="travel_dprd" if self.mode == "dprd" else "travel_secretariat",
            title=data.subject.strip() or "Draft Perjalanan Dinas",
            document_number=primary,
            document_date=data.letter_date.isoformat(),
            event_start=data.start_date.isoformat(),
            event_end=data.end_date.isoformat(),
            destination=" / ".join(data.destinations),
            status=status,
            author=self.user.username,
            payload=data.to_payload(),
            files=files or [],
            numbers=data.document_numbers if status == "generated" else {},
            record_id=self.record_id,
        )
        self.state_label.setText("Tersimpan sebagai draft" if status == "draft" else "Dokumen berhasil dibuat")
        self.changed.emit()

    def save_draft(self) -> None:
        try:
            self._save(self.collect(), "draft")
        except Exception as exc:
            QMessageBox.warning(self, "Draft gagal disimpan", str(exc))
            return
        self.audit.log(self.user.username, "sips_save_draft", self.state_label.text())
        QMessageBox.information(self, "Draft tersimpan", "Formulir perjalanan dinas telah disimpan sebagai draft.")

    def generate(self) -> None:
        output = QFileDialog.getExistingDirectory(self, "Pilih folder hasil dokumen")
        if not output:
            return
        try:
            data = self.collect()
            data.validate()
            self.repository.validate_numbers(data.document_numbers, self.record_id)
            QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
            QApplication.processEvents()
            report = self.service.generate_travel_report(data, output)
            files = report.files
            if not files:
                raise RuntimeError(report.error_message or "Tidak ada dokumen yang berhasil dibuat.")
            self._save(data, "generated", files)
        except Exception as exc:
            QMessageBox.critical(self, "Pembuatan dokumen gagal", str(exc))
            return
        finally:
            QApplication.restoreOverrideCursor()
        self.audit.log(self.user.username, "sips_generate_travel", f"{len(files)} file · {data.subject}")
        if report.failures:
            QMessageBox.warning(
                self,
                "Sebagian dokumen gagal dibuat",
                f"Berhasil membuat {len(files)} file di:\n{output}\n\n"
                f"Dokumen yang gagal:\n{report.error_message}",
            )
        else:
            QMessageBox.information(
                self,
                "Dokumen selesai dibuat",
                f"Berhasil membuat {len(files)} file di:\n{output}",
            )
        _open_path(output)

    def load_record(self, record_id: str) -> None:
        record = self.repository.get(record_id)
        if not record or not record.record_type.startswith("travel_"):
            return
        data = TravelFormData.from_payload(record.payload)
        if data.mode != self.mode:
            return
        self.record_id = record.record_id
        for key, field in self.number_fields.items():
            field.setText(data.document_numbers.get(key, ""))
        _set_date(self.letter_date, data.letter_date)
        _set_date(self.start_date, data.start_date)
        _set_date(self.end_date, data.end_date)
        self.travel_type.setCurrentText(data.travel_type)
        self.basis_dprd.setPlainText(data.basis_dprd)
        self.basis_asn.setPlainText(data.basis_asn)
        self.subject.setPlainText(data.subject)
        self.notice_subject.setPlainText(data.notice_subject)
        self.destinations.clear()
        self.destinations.addItems(data.destinations)
        self.signer_dprd.setCurrentText(data.signer_dprd)
        self.signer_asn.setCurrentText(data.signer_asn)
        if self.dprd_list:
            self.dprd_list.set_selected(data.dprd)
            self.asn_list.set_selected(data.asn)
        else:
            self.executor_list.set_selected(data.executors)
            self.companion_list.set_selected(data.companions)
        self.state_label.setText(f"Memuat {record.status}: {record.document_number or record.title}")
        self._update_duration()

    def reset_form(self) -> None:
        self.record_id = None
        for field in getattr(self, "number_fields", {}).values():
            field.clear()
        today = date.today()
        for widget in (getattr(self, "letter_date", None), getattr(self, "start_date", None), getattr(self, "end_date", None)):
            if widget:
                _set_date(widget, today)
        if hasattr(self, "basis_dprd"):
            self.basis_dprd.setPlainText("Keputusan Pimpinan DPRD Kota Bitung")
            self.basis_asn.setPlainText("Surat Perintah Sekretaris DPRD Kota Bitung")
            self.subject.setPlainText("")
            self.notice_subject.setPlainText("")
            self.destinations.clear()
            self.travel_type.setCurrentIndex(0)
            for listing in (self.dprd_list, self.asn_list, self.executor_list, self.companion_list):
                if listing:
                    listing.clear()
            self.state_label.setText("Formulir baru")
            self._update_duration()


class InvitationPage(QWidget):
    changed = Signal()

    def __init__(
        self,
        invitation_type: str,
        service: SIPSService,
        repository: SIPSRepository,
        user: User,
        audit: UserRepository,
    ) -> None:
        super().__init__()
        self.invitation_type = invitation_type
        self.service = service
        self.repository = repository
        self.user = user
        self.audit = audit
        self.record_id: str | None = None
        self._build_ui()
        self.reset_form()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 20)
        layout.setSpacing(14)
        title = QLabel("Undangan Paripurna" if self.invitation_type == "paripurna" else "Undangan Biasa")
        title.setObjectName("PageTitle")
        subtitle = QLabel("Buat undangan beserta Naskah Dinas dan Daftar Hadir dari template resmi kantor.")
        subtitle.setObjectName("Subtitle")
        layout.addWidget(title)
        layout.addWidget(subtitle)
        workspace = QSplitter(Qt.Orientation.Horizontal)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        host = QWidget()
        form = QGridLayout(host)
        form.setContentsMargins(4, 4, 12, 4)
        form.setHorizontalSpacing(16)
        form.setVerticalSpacing(12)
        self.number = QLineEdit()
        self.number.setPlaceholderText("Nomor awal undangan")
        self.letter_date = _date_edit()
        self.meeting_date = _date_edit()
        self.day_label = QLabel(day_name_id(date.today()))
        self.meeting_date.dateChanged.connect(lambda: self.day_label.setText(day_name_id(_python_date(self.meeting_date))))
        self.time_text = QLineEdit("09.00 WITA s.d. selesai")
        self.agenda = QPlainTextEdit()
        self.agenda.setMinimumHeight(100)
        self.signer = QComboBox()
        self.signer.setEditable(True)
        self.signer.addItems(self.service.master.dprd_signers)
        row = 0
        for label, widget in (
            ("Nomor undangan", self.number),
            ("Tanggal surat", self.letter_date),
            ("Tanggal rapat", self.meeting_date),
            ("Hari", self.day_label),
            ("Jam pelaksanaan", self.time_text),
            ("Isi surat / agenda", self.agenda),
            ("Penandatangan", self.signer),
        ):
            form.addWidget(QLabel(label), row, 0)
            form.addWidget(widget, row, 1)
            row += 1

        if self.invitation_type == "paripurna":
            self.clothing = QComboBox()
            self.clothing.addItems(("PSH", "PSR", "PSL", "PDU IV"))
            self.scenarios = QPlainTextEdit()
            self.scenarios.setPlaceholderText("Satu skenario per baris, maksimal tujuh skenario")
            self.scenarios.setMinimumHeight(110)
            form.addWidget(QLabel("Pakaian"), row, 0)
            form.addWidget(self.clothing, row, 1)
            row += 1
            form.addWidget(QLabel("Skenario rapat"), row, 0)
            form.addWidget(self.scenarios, row, 1)
            self.meeting_executor = self.meeting_type = self.related_parties = self.other_pages = None
        else:
            self.meeting_executor = QComboBox()
            self.meeting_executor.setEditable(True)
            self.meeting_executor.addItems(PELAKSANA_RAPAT_OPTIONS)
            self.meeting_executor.activated.connect(self._on_meeting_executor_selected)
            self.meeting_type = QComboBox()
            self.meeting_type.setEditable(True)
            self.meeting_type.addItems(JENIS_RAPAT_OPTIONS)
            self.related_parties = QPlainTextEdit()
            self.related_parties.setPlaceholderText("Satu pihak terkait per baris")
            self.related_parties.setMinimumHeight(100)
            self.other_pages = QPlainTextEdit()
            self.other_pages.setPlaceholderText(
                "Satu tujuan per baris. Pisahkan halaman tambahan dengan satu baris kosong."
            )
            self.other_pages.setMinimumHeight(110)
            for label, widget in (
                ("Pelaksana rapat", self.meeting_executor),
                ("Jenis rapat", self.meeting_type),
                ("Pihak terkait", self.related_parties),
                ("Tujuan surat lainnya", self.other_pages),
            ):
                form.addWidget(QLabel(label), row, 0)
                form.addWidget(widget, row, 1)
                row += 1
            self.clothing = self.scenarios = None

        support = QFrame()
        support.setObjectName("Card")
        support_layout = QVBoxLayout(support)
        support_layout.addWidget(QLabel("Dokumen Pendukung"))
        self.include_note = QCheckBox("Buat Naskah Dinas")
        self.include_attendance = QCheckBox("Buat Daftar Hadir")
        support_layout.addWidget(self.include_note)
        support_layout.addWidget(self.include_attendance)
        self.include_related = QCheckBox("Sertakan lembar Pihak Terkait")
        self.include_related.setChecked(True)
        self.include_secretariat = QCheckBox("Sertakan lembar Sekretariat")
        if self.invitation_type == "biasa":
            support_layout.addWidget(self.include_related)
            support_layout.addWidget(self.include_secretariat)
        else:
            self.include_related.setVisible(False)
            self.include_secretariat.setVisible(False)
        support_actions = QHBoxLayout()
        create_note = QPushButton("Buat Naskah Dinas Saja")
        create_attendance = QPushButton("Buat Daftar Hadir Saja")
        create_note.clicked.connect(lambda: self._generate_support_document("note"))
        create_attendance.clicked.connect(lambda: self._generate_support_document("attendance"))
        support_actions.addWidget(create_note)
        support_actions.addWidget(create_attendance)
        support_layout.addLayout(support_actions)
        form.addWidget(support, row + 1, 0, 1, 2)
        form.setColumnStretch(1, 1)
        scroll.setWidget(host)
        workspace.addWidget(scroll)
        self.live_preview = LiveDocumentPreview()
        workspace.addWidget(self.live_preview)
        workspace.setStretchFactor(0, 5)
        workspace.setStretchFactor(1, 3)
        workspace.setChildrenCollapsible(False)
        workspace.setSizes((760, 440))
        layout.addWidget(workspace, 1)
        actions = QHBoxLayout()
        self.state_label = QLabel("Formulir baru")
        actions.addWidget(self.state_label, 1)
        reset = QPushButton("Formulir Baru")
        draft = QPushButton("Simpan Draft")
        generate = QPushButton("Buat Dokumen")
        generate.setObjectName("PrimaryButton")
        reset.clicked.connect(self.reset_form)
        draft.clicked.connect(self.save_draft)
        generate.clicked.connect(self.generate)
        for button in (reset, draft, generate):
            actions.addWidget(button)
        layout.addLayout(actions)
        self._connect_live_preview_signals()

    def _connect_live_preview_signals(self) -> None:
        self.number.textChanged.connect(self._schedule_live_preview)
        self.time_text.textChanged.connect(self._schedule_live_preview)
        self.agenda.textChanged.connect(self._schedule_live_preview)
        self.letter_date.dateChanged.connect(self._schedule_live_preview)
        self.meeting_date.dateChanged.connect(self._schedule_live_preview)
        self.signer.currentTextChanged.connect(self._schedule_live_preview)
        for checkbox in (
            self.include_note,
            self.include_attendance,
            self.include_related,
            self.include_secretariat,
        ):
            checkbox.toggled.connect(self._schedule_live_preview)
        if self.clothing:
            self.clothing.currentTextChanged.connect(self._schedule_live_preview)
            self.scenarios.textChanged.connect(self._schedule_live_preview)
        else:
            self.meeting_executor.currentTextChanged.connect(self._schedule_live_preview)
            self.meeting_type.currentTextChanged.connect(self._schedule_live_preview)
            self.related_parties.textChanged.connect(self._schedule_live_preview)
            self.other_pages.textChanged.connect(self._schedule_live_preview)

    def _on_meeting_executor_selected(self, _index: int) -> None:
        if self.meeting_executor and self.meeting_executor.currentText() == PELAKSANA_RAPAT_CUSTOM:
            self.meeting_executor.setEditText("")
            self.meeting_executor.setFocus()

    def _schedule_live_preview(self, *_args) -> None:
        data = self.collect()
        try:
            data.validate_preview()
        except ValueError as exc:
            self.live_preview.show_waiting(f"Live preview menunggu: {exc}")
            return
        self.live_preview.schedule(
            lambda output, current=data: self.service.generate_invitation(current, output, preview=True)
        )

    def _generate_support_document(self, kind: str) -> None:
        title = "Naskah Dinas" if kind == "note" else "Daftar Hadir"
        output = QFileDialog.getExistingDirectory(self, f"Pilih folder hasil {title}")
        if not output:
            return
        try:
            data = self.collect()
            QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
            QApplication.processEvents()
            if kind == "note":
                path = self.service.generate_official_note(data, output)
                action = "sips_generate_official_note"
            else:
                path = self.service.generate_meeting_attendance(data, output)
                action = "sips_generate_meeting_attendance"
        except Exception as exc:
            QMessageBox.critical(self, f"Pembuatan {title} gagal", str(exc))
            return
        finally:
            QApplication.restoreOverrideCursor()
        self.audit.log(self.user.username, action, f"{self.invitation_type} · {path.name}")
        QMessageBox.information(self, f"{title} selesai dibuat", f"File tersimpan di:\n{path}")
        _open_path(path)

    def _parse_pages(self) -> list[list[str]]:
        if not self.other_pages:
            return []
        text = self.other_pages.toPlainText().strip()
        if not text:
            return []
        return [
            [line.strip() for line in block.splitlines() if line.strip()]
            for block in re.split(r"\n\s*\n", text)
            if block.strip()
        ]

    def collect(self) -> InvitationFormData:
        return InvitationFormData(
            invitation_type=self.invitation_type,
            number=self.number.text().strip(),
            letter_date=_python_date(self.letter_date),
            meeting_date=_python_date(self.meeting_date),
            time_text=self.time_text.text().strip(),
            agenda=self.agenda.toPlainText(),
            signer=self.signer.currentText(),
            clothing=self.clothing.currentText() if self.clothing else "",
            scenarios=_lines(self.scenarios)[:7] if self.scenarios else [],
            meeting_executor=self.meeting_executor.currentText() if self.meeting_executor else "",
            meeting_type=self.meeting_type.currentText() if self.meeting_type else "",
            related_parties=_lines(self.related_parties) if self.related_parties else [],
            other_destination_pages=self._parse_pages(),
            include_official_note=self.include_note.isChecked(),
            include_attendance=self.include_attendance.isChecked(),
            include_related_attendance=self.include_related.isChecked(),
            include_secretariat_attendance=self.include_secretariat.isChecked(),
        )

    def _save(self, data: InvitationFormData, status: str, files: list[Path] = None) -> None:
        self.record_id = self.repository.save(
            record_type="invitation_plenary" if self.invitation_type == "paripurna" else "invitation_regular",
            title=data.agenda.strip() or "Draft Surat Undangan",
            document_number=data.number,
            document_date=data.letter_date.isoformat(),
            event_start=data.meeting_date.isoformat(),
            event_end=data.meeting_date.isoformat(),
            destination=data.meeting_executor or "Pimpinan dan Anggota DPRD Kota Bitung",
            status=status,
            author=self.user.username,
            payload=data.to_payload(),
            files=files or [],
            numbers={"nomor_undangan": data.number} if status == "generated" else {},
            record_id=self.record_id,
        )
        self.state_label.setText("Tersimpan sebagai draft" if status == "draft" else "Dokumen berhasil dibuat")
        self.changed.emit()

    def save_draft(self) -> None:
        try:
            self._save(self.collect(), "draft")
        except Exception as exc:
            QMessageBox.warning(self, "Draft gagal disimpan", str(exc))
            return
        self.audit.log(self.user.username, "sips_save_draft", f"Undangan {self.invitation_type}")
        QMessageBox.information(self, "Draft tersimpan", "Formulir undangan telah disimpan sebagai draft.")

    def generate(self) -> None:
        output = QFileDialog.getExistingDirectory(self, "Pilih folder hasil dokumen")
        if not output:
            return
        try:
            data = self.collect()
            data.validate()
            self.repository.validate_numbers({"nomor_undangan": data.number}, self.record_id)
            QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
            QApplication.processEvents()
            report = self.service.generate_invitation_report(data, output)
            files = report.files
            if not files:
                raise RuntimeError(report.error_message or "Tidak ada dokumen yang berhasil dibuat.")
            self._save(data, "generated", files)
        except Exception as exc:
            QMessageBox.critical(self, "Pembuatan dokumen gagal", str(exc))
            return
        finally:
            QApplication.restoreOverrideCursor()
        self.audit.log(self.user.username, "sips_generate_invitation", f"{data.invitation_type} · {data.number}")
        if report.failures:
            QMessageBox.warning(
                self,
                "Sebagian dokumen gagal dibuat",
                f"Berhasil membuat {len(files)} file di:\n{output}\n\n"
                f"Dokumen yang gagal:\n{report.error_message}",
            )
        else:
            QMessageBox.information(self, "Dokumen selesai dibuat", f"Berhasil membuat {len(files)} file di:\n{output}")
        _open_path(output)

    def load_record(self, record_id: str) -> None:
        record = self.repository.get(record_id)
        if not record or not record.record_type.startswith("invitation_"):
            return
        data = InvitationFormData.from_payload(record.payload)
        if data.invitation_type != self.invitation_type:
            return
        self.record_id = record.record_id
        self.number.setText(data.number)
        _set_date(self.letter_date, data.letter_date)
        _set_date(self.meeting_date, data.meeting_date)
        self.time_text.setText(data.time_text)
        self.agenda.setPlainText(data.agenda)
        self.signer.setCurrentText(data.signer)
        if self.clothing:
            self.clothing.setCurrentText(data.clothing)
            self.scenarios.setPlainText("\n".join(data.scenarios))
        else:
            self.meeting_executor.setCurrentText(data.meeting_executor)
            self.meeting_type.setCurrentText(data.meeting_type)
            self.related_parties.setPlainText("\n".join(data.related_parties))
            self.other_pages.setPlainText("\n\n".join("\n".join(page) for page in data.other_destination_pages))
        self.include_note.setChecked(data.include_official_note)
        self.include_attendance.setChecked(data.include_attendance)
        self.include_related.setChecked(data.include_related_attendance)
        self.include_secretariat.setChecked(data.include_secretariat_attendance)
        self.day_label.setText(day_name_id(data.meeting_date))
        self.state_label.setText(f"Memuat {record.status}: {record.document_number}")

    def reset_form(self) -> None:
        self.record_id = None
        if not hasattr(self, "number"):
            return
        self.number.clear()
        today = date.today()
        _set_date(self.letter_date, today)
        _set_date(self.meeting_date, today)
        self.time_text.setText("09.00 WITA s.d. selesai")
        self.agenda.clear()
        if self.clothing:
            self.clothing.setCurrentText("PSH")
            self.scenarios.clear()
        else:
            self.meeting_executor.setCurrentIndex(0)
            self.meeting_type.setCurrentIndex(0)
            self.related_parties.clear()
            self.other_pages.clear()
        self.include_note.setChecked(False)
        self.include_attendance.setChecked(False)
        self.include_related.setChecked(True)
        self.include_secretariat.setChecked(False)
        self.day_label.setText(day_name_id(today))
        self.state_label.setText("Formulir baru")


class SIPSRecapPage(QWidget):
    edit_requested = Signal(str, str)

    def __init__(self, category: str, repository: SIPSRepository):
        super().__init__()
        self.category = category
        self.repository = repository
        self.records: list[DocumentRecord] = []
        self._build_ui()
        self.refresh()

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setContentsMargins(28, 24, 28, 22)
        layout.setSpacing(14)
        title = QLabel("Rekapitulasi Surat Perjalanan Dinas" if self.category == "travel" else "Rekapitulasi Surat Undangan")
        title.setObjectName("PageTitle")
        subtitle = QLabel("Riwayat dokumen dan draft tersimpan pada database lokal terpadu.")
        subtitle.setObjectName("Subtitle")
        layout.addWidget(title)
        layout.addWidget(subtitle)
        controls = QHBoxLayout()
        self.search = QLineEdit()
        self.search.setPlaceholderText("Cari nomor, agenda, tujuan, atau pembuat…")
        self.search.setClearButtonEnabled(True)
        self.search.textChanged.connect(self.refresh)
        self.status = QComboBox()
        self.status.addItems(("Semua Status", "Generated", "Draft"))
        self.status.currentTextChanged.connect(self.refresh)
        export = QPushButton("Ekspor Rekap Excel")
        export.setObjectName("PrimaryButton")
        export.clicked.connect(self.export_excel)
        controls.addWidget(self.search, 1)
        controls.addWidget(self.status)
        controls.addWidget(export)
        layout.addLayout(controls)
        self.table = QTableWidget(0, 9)
        self.table.setHorizontalHeaderLabels(("Tanggal", "Nomor Surat", "Jenis", "Agenda", "Tujuan/Pelaksana", "Status", "Pembuat", "Diperbarui", "ID"))
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.verticalHeader().setVisible(False)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.table.setColumnHidden(8, True)
        self.table.doubleClicked.connect(self.edit_selected)
        layout.addWidget(self.table, 1)
        actions = QHBoxLayout()
        edit = QPushButton("Edit/Muat Formulir")
        open_doc = QPushButton("Buka Dokumen")
        open_folder = QPushButton("Buka Folder")
        delete = QPushButton("Hapus Draft")
        self.printers = QComboBox()
        self.printers.setMinimumWidth(220)
        self.printers.addItems(QPrinterInfo.availablePrinterNames() or ["Printer tidak ditemukan"])
        print_button = QPushButton("Cetak Dokumen")
        edit.clicked.connect(self.edit_selected)
        open_doc.clicked.connect(self.open_document)
        open_folder.clicked.connect(self.open_folder)
        delete.clicked.connect(self.delete_draft)
        print_button.clicked.connect(self.print_document)
        for widget in (edit, open_doc, open_folder, delete):
            actions.addWidget(widget)
        actions.addStretch()
        actions.addWidget(self.printers)
        actions.addWidget(print_button)
        layout.addLayout(actions)

    def refresh(self, *_args) -> None:
        records = self.repository.list(category=self.category, search=self.search.text() if hasattr(self, "search") else "")
        if hasattr(self, "status") and self.status.currentText() != "Semua Status":
            wanted = self.status.currentText().lower()
            records = [record for record in records if record.status == wanted]
        self.records = records
        self.table.setRowCount(len(records))
        labels = {
            "travel_dprd": "Perjalanan DPRD",
            "travel_secretariat": "Perjalanan Setwan",
            "invitation_plenary": "Undangan Paripurna",
            "invitation_regular": "Undangan Biasa",
        }
        for row, record in enumerate(records):
            values = (
                record.document_date,
                record.document_number,
                labels.get(record.record_type, record.record_type),
                record.title,
                record.destination,
                "DRAFT" if record.status == "draft" else "SELESAI",
                record.author,
                record.updated_at.replace("T", " "),
                record.record_id,
            )
            for column, value in enumerate(values):
                item = QTableWidgetItem(str(value))
                if column == 5 and record.status == "draft":
                    item.setBackground(Qt.GlobalColor.yellow)
                self.table.setItem(row, column, item)
        self.table.resizeColumnsToContents()
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self.table.setColumnHidden(8, True)

    def _selected(self) -> DocumentRecord | None:
        row = self.table.currentRow()
        if row < 0:
            QMessageBox.information(self, "Belum ada pilihan", "Pilih satu baris dokumen terlebih dahulu.")
            return None
        record_id = self.table.item(row, 8).text()
        return next((record for record in self.records if record.record_id == record_id), None)

    def edit_selected(self, *_args) -> None:
        record = self._selected()
        if not record:
            return
        route = {
            "travel_dprd": "travel_dprd",
            "travel_secretariat": "travel_secretariat",
            "invitation_plenary": "invitation_plenary",
            "invitation_regular": "invitation_regular",
        }[record.record_type]
        self.edit_requested.emit(route, record.record_id)

    def open_document(self) -> None:
        record = self._selected()
        if not record:
            return
        path = self._choose_document(record, "Dokumen yang akan dibuka")
        if not path:
            return
        _open_path(path)

    def open_folder(self) -> None:
        record = self._selected()
        if not record:
            return
        path = next((Path(value) for value in record.files if Path(value).exists()), None)
        if path:
            _open_path(path.parent)
        else:
            QMessageBox.warning(self, "Folder tidak ditemukan", "Lokasi hasil dokumen tidak tersedia.")

    def print_document(self) -> None:
        record = self._selected()
        if not record:
            return
        path = self._choose_document(record, "Dokumen yang akan dicetak")
        if not path:
            return
        try:
            _print_to(path, self.printers.currentText() if self.printers.count() else "")
        except Exception as exc:
            QMessageBox.critical(self, "Pencetakan gagal", str(exc))
            return
        QMessageBox.information(self, "Pencetakan", f"Dokumen dikirim ke {self.printers.currentText()}.")

    def _choose_document(self, record: DocumentRecord, title: str) -> Path | None:
        available = [Path(value) for value in record.files if Path(value).exists()]
        if not available:
            QMessageBox.warning(
                self,
                "File tidak ditemukan",
                "Dokumen belum dibuat atau telah dipindahkan dari lokasi asal.",
            )
            return None
        if len(available) == 1:
            return available[0]
        selected, accepted = QInputDialog.getItem(
            self,
            "Pilih dokumen",
            f"{title}:",
            [path.name for path in available],
            0,
            False,
        )
        if not accepted:
            return None
        return next(path for path in available if path.name == selected)

    def delete_draft(self) -> None:
        record = self._selected()
        if not record:
            return
        if record.status != "draft":
            QMessageBox.warning(self, "Tidak dapat dihapus", "Hanya draft yang dapat dihapus dari halaman ini.")
            return
        if QMessageBox.question(self, "Hapus draft", "Hapus draft terpilih?") != QMessageBox.StandardButton.Yes:
            return
        self.repository.delete_draft(record.record_id)
        self.refresh()

    def export_excel(self) -> None:
        suggested = "Rekap_Perjalanan_Dinas.xlsx" if self.category == "travel" else "Rekap_Surat_Undangan.xlsx"
        path, _ = QFileDialog.getSaveFileName(self, "Ekspor rekap", suggested, "Excel Workbook (*.xlsx)")
        if not path:
            return
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Rekap"
        headers = ("No", "Tanggal Surat", "Nomor Surat", "Jenis", "Agenda/Perihal", "Tujuan/Pelaksana", "Status", "Pembuat", "Diperbarui", "Lokasi File")
        for column, value in enumerate(headers, start=1):
            cell = sheet.cell(1, column, value)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1E3A8A")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        medium = Side(style="medium", color="000000")
        border = Border(left=medium, right=medium, top=medium, bottom=medium)
        labels = {
            "travel_dprd": "Perjalanan DPRD", "travel_secretariat": "Perjalanan Setwan",
            "invitation_plenary": "Undangan Paripurna", "invitation_regular": "Undangan Biasa",
        }
        for row, record in enumerate(self.records, start=2):
            first_file = record.files[0] if record.files else ""
            values = (row - 1, record.document_date, record.document_number, labels.get(record.record_type, record.record_type), record.title, record.destination, record.status.upper(), record.author, record.updated_at.replace("T", " "), first_file)
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row, column, value)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
        for cell in sheet[1]:
            cell.border = border
        widths = (7, 15, 24, 22, 46, 34, 12, 16, 21, 55)
        for column, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        workbook.save(path)
        QMessageBox.information(self, "Ekspor selesai", f"Rekap Excel tersimpan di:\n{path}")
