from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, timedelta
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.table import Table, TableStyleInfo

from tpp_finger_scan.application.services import RecapSession
from tpp_finger_scan.domain.calendar import day_name_id, schedule_for
from tpp_finger_scan.domain.models import DayCalculation, SpecialCode, ZERO


NAVY = "17324D"
BLUE = "1D73E8"
LIGHT_BLUE = "EAF2FF"
YELLOW = "FFF2CC"
LIGHT_RED = "FCE8E6"
WHITE = "FFFFFF"
GRAY = "667085"
THIN_GRAY = Side(style="thin", color="D0D5DD")
# Rekap cetak mengikuti lampiran: seluruh garis tabel utama memakai ketebalan
# medium agar tidak bercampur antara garis tipis dan tebal.
THIN_BLACK = Side(style="medium", color="000000")
MEDIUM_BLACK = Side(style="medium", color="000000")
PERCENT_FORMAT = "0.00%"
MONTH_NAMES_ID = (
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
)


@dataclass(frozen=True, slots=True)
class RecapExportProfile:
    signer_title: str = "Kepala Bagian Umum dan Keuangan"
    signer_name: str = "SANTY N. MAMESAH, SS, M.Si"
    signer_nip: str = "198109112003122005"
    exclude_source_boundary_dates: bool = True


class ExcelExporter:
    def __init__(self, profile: RecapExportProfile | None = None) -> None:
        self.profile = profile or RecapExportProfile()

    def export(self, session: RecapSession, destination: str | Path) -> Path:
        target = Path(destination).expanduser().resolve()
        if target.suffix.lower() != ".xlsx":
            target = target.with_suffix(".xlsx")
        target.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        employee_recap = workbook.active
        employee_recap.title = "Rekap Per Pegawai"
        master = workbook.create_sheet("Master Pegawai")
        master_rows = self._build_employee_master(master, session)
        self._build_employee_recap(employee_recap, session, master_rows)
        summary = workbook.create_sheet("Ringkasan")
        self._build_summary(summary, session)
        detail = workbook.create_sheet("Detail Harian")
        self._build_detail(detail, session)
        rules = workbook.create_sheet("Catatan Aturan")
        self._build_rule_notes(rules)

        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.save(target)
        return target

    def _build_employee_master(self, sheet, session: RecapSession) -> dict[str, int]:
        sheet.sheet_view.showGridLines = False
        headers = ("ID Finger", "Nama Pegawai", "Jabatan")
        for column, value in enumerate(headers, start=1):
            self._header_style(sheet.cell(1, column, value))
        row_map: dict[str, int] = {}
        for row_index, employee in enumerate(session.import_result.employees, start=2):
            row_map[employee.finger_id] = row_index
            values = (
                employee.finger_id,
                employee.name,
                session.employee_positions.get(employee.finger_id, ""),
            )
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row_index, column, value)
                self._body_style(cell)
            sheet.cell(row_index, 1).number_format = "@"
        last_row = len(session.import_result.employees) + 1
        if last_row > 1:
            table = Table(displayName="MasterPegawai", ref=f"A1:C{last_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
        sheet.column_dimensions["A"].width = 15
        sheet.column_dimensions["B"].width = 34
        sheet.column_dimensions["C"].width = 42
        sheet.freeze_panes = "A2"
        sheet["E2"] = "Petunjuk"
        sheet["E2"].font = Font(bold=True, color=NAVY)
        sheet["E3"] = (
            "Isi atau perbaiki kolom Jabatan. Sheet Rekap Per Pegawai mengambil "
            "jabatan dari tabel ini secara otomatis."
        )
        sheet["E3"].alignment = Alignment(wrap_text=True, vertical="top")
        sheet.column_dimensions["E"].width = 46
        return row_map

    def _build_employee_recap(
        self,
        sheet,
        session: RecapSession,
        master_rows: dict[str, int],
    ) -> None:
        sheet.sheet_view.showGridLines = False
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.page_margins.left = 0.2
        sheet.page_margins.right = 0.2
        sheet.page_margins.top = 0.25
        sheet.page_margins.bottom = 0.25
        sheet.page_margins.header = 0.1
        sheet.page_margins.footer = 0.1
        sheet.sheet_view.zoomScale = 80

        widths = (13, 14, 13, 13, 13, 15, 15, 15, 15, 15, 13)
        for column, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(column)].width = width

        calculations = {
            (item.entry.employee.finger_id, item.entry.work_date): item
            for item in session.calculations
        }
        report_dates = self._report_dates(session)
        report_period = self._format_period_id(report_dates)
        current_row = 1
        last_employee_index = len(session.import_result.employees) - 1

        for employee_index, employee in enumerate(session.import_result.employees):
            start_row = current_row
            title_row = start_row
            name_row = start_row + 1
            position_row = start_row + 2
            period_row = start_row + 3
            group_header_row = start_row + 5
            sub_header_row = start_row + 6
            number_row = start_row + 7
            separator_row = start_row + 8
            data_start_row = start_row + 9
            data_end_row = data_start_row + len(report_dates) - 1
            total_row = data_end_row + 1
            signer_title_row = total_row + 2
            signer_name_row = total_row + 5
            signer_nip_row = total_row + 6
            block_end_row = total_row + 8

            sheet.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=11)
            title_cell = sheet.cell(title_row, 1)
            title_cell.value = (
                "FORMAT REKAPITULASI PEMOTONGAN DISIPLIN KERJA DAN "
                "PRESTASI KERJA SETIAP PNS"
            )
            title_cell.font = Font(name="Arial", size=10, bold=True)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            title_cell.border = Border(top=MEDIUM_BLACK, left=MEDIUM_BLACK, right=MEDIUM_BLACK)
            sheet.row_dimensions[title_row].height = 17

            metadata = (("NAMA", employee.name), ("JABATAN", None), ("BULAN", report_period))
            master_row = master_rows[employee.finger_id]
            for row_index, (label, value) in zip(
                (name_row, position_row, period_row), metadata, strict=True,
            ):
                sheet.cell(row_index, 1, label)
                sheet.merge_cells(start_row=row_index, start_column=2, end_row=row_index, end_column=11)
                value_cell = sheet.cell(row_index, 2)
                if row_index == position_row:
                    value_cell.value = (
                        f'=IF(\'Master Pegawai\'!C{master_row}="",": -",'
                        f'": "&\'Master Pegawai\'!C{master_row})'
                    )
                else:
                    value_cell.value = f": {value}"
                for column in range(1, 12):
                    cell = sheet.cell(row_index, column)
                    cell.font = Font(name="Arial", size=9)
                    cell.alignment = Alignment(vertical="center")
                sheet.row_dimensions[row_index].height = 15

            sheet.merge_cells(
                start_row=group_header_row,
                start_column=1,
                end_row=sub_header_row,
                end_column=1,
            )
            sheet.cell(group_header_row, 1, "Tanggal")
            sheet.merge_cells(
                start_row=group_header_row,
                start_column=2,
                end_row=group_header_row,
                end_column=5,
            )
            sheet.cell(group_header_row, 2, "Disiplin Kerja")
            sheet.merge_cells(
                start_row=group_header_row,
                start_column=6,
                end_row=group_header_row,
                end_column=11,
            )
            sheet.cell(group_header_row, 6, "Produktivitas Kerja")

            sub_headers = (
                "Tidak masuk kerja",
                "Terlambat",
                "Pulang Cepat",
                "Jumlah",
                "Nilai Laporan\nProduktivitas\nKerja 75–100",
                "Nilai Laporan\nProduktivitas\nKerja 50–<75",
                "Nilai Laporan\nProduktivitas\nKerja 25–<50",
                "Nilai Laporan\nProduktivitas\nKerja 1–<25",
                "Nilai Laporan\nProduktivitas\nKerja 0/Tidak\nMembuat Lap.",
                "Jumlah",
            )
            for column, value in enumerate(sub_headers, start=2):
                sheet.cell(sub_header_row, column, value)
            for row_index in (group_header_row, sub_header_row):
                for column in range(1, 12):
                    cell = sheet.cell(row_index, column)
                    cell.font = Font(name="Arial", size=7, bold=False)
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = Border(
                        left=MEDIUM_BLACK if column in {1, 2, 6} else THIN_BLACK,
                        right=MEDIUM_BLACK if column in {1, 5, 11} else THIN_BLACK,
                        top=MEDIUM_BLACK if row_index == group_header_row else THIN_BLACK,
                        bottom=MEDIUM_BLACK if row_index == sub_header_row else THIN_BLACK,
                    )
            sheet.row_dimensions[group_header_row].height = 16
            sheet.row_dimensions[sub_header_row].height = 51

            for column in range(1, 12):
                cell = sheet.cell(number_row, column, column)
                cell.font = Font(name="Arial", size=7)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=THIN_BLACK, right=THIN_BLACK, bottom=MEDIUM_BLACK)
            sheet.row_dimensions[number_row].height = 15
            for column in range(1, 12):
                separator = sheet.cell(separator_row, column)
                separator.fill = PatternFill("solid", fgColor="A6A6A6")
                separator.border = Border(left=THIN_BLACK, right=THIN_BLACK, bottom=MEDIUM_BLACK)
            sheet.row_dimensions[separator_row].height = 9

            for offset, work_date in enumerate(report_dates):
                row_index = data_start_row + offset
                calculation = calculations.get((employee.finger_id, work_date))
                sheet.cell(row_index, 1, work_date)
                sheet.cell(row_index, 1).number_format = "dd/mm/yyyy"
                if calculation is not None:
                    override = session.overrides.get((employee.finger_id, work_date))
                    if override and override.code in {SpecialCode.TL, SpecialCode.WFH, SpecialCode.W, SpecialCode.S}:
                        sheet.cell(row_index, 2, override.code.value)
                        if override.code == SpecialCode.S:
                            for column in range(1, 6):
                                sheet.cell(row_index, column).fill = PatternFill("solid", fgColor=YELLOW)
                    elif calculation.deductions.absence > ZERO:
                        sheet.cell(row_index, 2, self._excel_percent(calculation.deductions.absence))
                    if calculation.deductions.late > ZERO:
                        sheet.cell(row_index, 3, self._excel_percent(calculation.deductions.late))
                    if calculation.deductions.early > ZERO:
                        sheet.cell(row_index, 4, self._excel_percent(calculation.deductions.early))
                sheet.cell(row_index, 5, f"=SUM(B{row_index}:D{row_index})")
                for column in range(1, 12):
                    cell = sheet.cell(row_index, column)
                    cell.font = Font(name="Arial", size=8)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(left=THIN_BLACK, right=THIN_BLACK)
                for column in range(2, 12):
                    sheet.cell(row_index, column).number_format = PERCENT_FORMAT
                sheet.row_dimensions[row_index].height = 15

            sheet.cell(total_row, 1, "Jumlah")
            for column in range(2, 6):
                letter = get_column_letter(column)
                sheet.cell(total_row, column, f"=SUM({letter}{data_start_row}:{letter}{data_end_row})")
                sheet.cell(total_row, column).number_format = PERCENT_FORMAT
            for column in range(6, 12):
                sheet.cell(total_row, column, "")
            for column in range(1, 12):
                cell = sheet.cell(total_row, column)
                cell.font = Font(name="Arial", size=8, bold=True if column == 1 else False)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(
                    left=MEDIUM_BLACK if column in {1, 2, 6} else THIN_BLACK,
                    right=MEDIUM_BLACK if column in {1, 5, 11} else THIN_BLACK,
                    top=MEDIUM_BLACK,
                    bottom=MEDIUM_BLACK,
                )
            sheet.row_dimensions[total_row].height = 16

            for row_index, value, underlined in (
                (signer_title_row, self.profile.signer_title, False),
                (signer_name_row, self.profile.signer_name, True),
                (signer_nip_row, f"NIP. {self.profile.signer_nip}", False),
            ):
                sheet.merge_cells(start_row=row_index, start_column=7, end_row=row_index, end_column=11)
                cell = sheet.cell(row_index, 7, value)
                cell.font = Font(name="Arial", size=9, underline="single" if underlined else None)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for column in range(1, 12):
                sheet.cell(block_end_row, column).border = Border(bottom=MEDIUM_BLACK)
            if employee_index < last_employee_index:
                sheet.row_breaks.append(Break(id=block_end_row))
            current_row = block_end_row + 1

        if current_row > 1:
            sheet.print_area = f"A1:K{current_row - 1}"

    def _report_dates(self, session: RecapSession) -> list[date]:
        start = session.import_result.period_start
        end = session.import_result.period_end
        if self.profile.exclude_source_boundary_dates and (end - start).days >= 2:
            start += timedelta(days=1)
            end -= timedelta(days=1)
        dates = []
        current = start
        while current <= end:
            if schedule_for(current).workday and current not in session.holidays:
                dates.append(current)
            current += timedelta(days=1)
        return dates

    @staticmethod
    def _format_period_id(dates: list[date]) -> str:
        if not dates:
            return "-"
        start, end = dates[0], dates[-1]
        if start.year != end.year:
            return (
                f"{start.day} {MONTH_NAMES_ID[start.month - 1]} {start.year} - "
                f"{end.day} {MONTH_NAMES_ID[end.month - 1]} {end.year}"
            )
        if start.month == end.month:
            return f"{start.day} - {end.day} {MONTH_NAMES_ID[end.month - 1]} {end.year}"
        return (
            f"{start.day} {MONTH_NAMES_ID[start.month - 1]} - "
            f"{end.day} {MONTH_NAMES_ID[end.month - 1]} {end.year}"
        )

    def _build_summary(self, sheet, session: RecapSession) -> None:
        result = session.import_result
        report_dates = self._report_dates(session)
        report_date_set = set(report_dates)
        sheet.sheet_view.showGridLines = False
        sheet.merge_cells("A1:H1")
        sheet["A1"] = "REKAPITULASI POTONGAN DISIPLIN KERJA"
        sheet["A1"].font = Font(size=16, bold=True, color=WHITE)
        sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
        sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 30
        sheet["A3"] = "Periode Rekap"
        sheet["B3"] = self._format_period_id(report_dates)
        sheet["A4"] = "Sumber"
        sheet["B4"] = result.source_path.name
        sheet["A5"] = "SHA-256"
        sheet["B5"] = result.source_sha256
        sheet.merge_cells("B5:H5")

        headers = [
            "No", "ID Finger", "Nama Pegawai", "Tidak Masuk", "Terlambat",
            "Pulang Cepat", "Jumlah Potongan", "Perlu Review",
        ]
        header_row = 7
        for column, value in enumerate(headers, start=1):
            cell = sheet.cell(header_row, column, value)
            self._header_style(cell)

        grouped: dict[str, list[DayCalculation]] = defaultdict(list)
        for calculation in session.calculations:
            if calculation.entry.work_date in report_date_set:
                grouped[calculation.entry.employee.finger_id].append(calculation)

        for row_index, employee in enumerate(result.employees, start=header_row + 1):
            calculations = grouped[employee.finger_id]
            absence = sum((item.deductions.absence for item in calculations), ZERO)
            late = sum((item.deductions.late for item in calculations), ZERO)
            early = sum((item.deductions.early for item in calculations), ZERO)
            total = absence + late + early
            values = [
                row_index - header_row,
                employee.finger_id,
                employee.name,
                self._excel_percent(absence),
                self._excel_percent(late),
                self._excel_percent(early),
                self._excel_percent(total),
                sum(not item.finalizable for item in calculations),
            ]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row_index, column, value)
                self._body_style(cell)
                if column in {4, 5, 6, 7}:
                    cell.number_format = PERCENT_FORMAT

        end_row = header_row + len(result.employees)
        if end_row > header_row:
            table = Table(displayName="RingkasanPegawai", ref=f"A{header_row}:H{end_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
        widths = [16, 14, 32, 16, 14, 16, 18, 15]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.auto_filter.ref = f"A{header_row}:H{end_row}"

    def _build_detail(self, sheet, session: RecapSession) -> None:
        sheet.sheet_view.showGridLines = False
        headers = [
            "No", "Tanggal", "Hari", "ID Finger", "Nama Pegawai", "Masuk",
            "Pulang", "Kode", "Rawat Inap", "Tidak Masuk", "Terlambat",
            "Pulang Cepat", "Jumlah", "Status", "Catatan Review", "Halaman PDF",
        ]
        for column, value in enumerate(headers, start=1):
            self._header_style(sheet.cell(1, column, value))

        for row_index, calculation in enumerate(session.calculations, start=2):
            entry = calculation.entry
            key = (entry.employee.finger_id, entry.work_date)
            override = session.overrides.get(key)
            values = [
                row_index - 1,
                entry.work_date,
                day_name_id(entry.work_date),
                entry.employee.finger_id,
                entry.employee.name,
                entry.in_time.strftime("%H:%M") if entry.in_time else "-",
                entry.out_time.strftime("%H:%M") if entry.out_time else "-",
                override.code.value if override else "",
                "Ya" if override and override.inpatient else "",
                self._excel_percent(calculation.deductions.absence),
                self._excel_percent(calculation.deductions.late),
                self._excel_percent(calculation.deductions.early),
                self._excel_percent(calculation.deductions.total),
                calculation.status,
                " | ".join(issue.message for issue in calculation.issues),
                entry.source_page,
            ]
            for column, value in enumerate(values, start=1):
                cell = sheet.cell(row_index, column, value)
                self._body_style(cell)
                if column == 2:
                    cell.number_format = "dd/mm/yyyy"
                elif column in {10, 11, 12, 13}:
                    cell.number_format = PERCENT_FORMAT
            if calculation.highlight_yellow:
                for column in range(1, len(headers) + 1):
                    sheet.cell(row_index, column).fill = PatternFill("solid", fgColor=YELLOW)
            elif not calculation.finalizable:
                for column in range(1, len(headers) + 1):
                    sheet.cell(row_index, column).fill = PatternFill("solid", fgColor=LIGHT_RED)

        end_row = len(session.calculations) + 1
        if end_row > 1:
            table = Table(displayName="DetailKehadiran", ref=f"A1:P{end_row}")
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            sheet.add_table(table)
        widths = [7, 13, 12, 14, 30, 10, 10, 9, 12, 15, 13, 15, 12, 24, 55, 13]
        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:P{end_row}"
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True

    @staticmethod
    def _build_rule_notes(sheet) -> None:
        sheet.sheet_view.showGridLines = False
        sheet["A1"] = "CATATAN KONFIGURASI ATURAN"
        sheet["A1"].font = Font(size=15, bold=True, color=WHITE)
        sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
        sheet.merge_cells("A1:C1")
        rows = [
            ("Masuk", "07.40–08.00", "0,50%"),
            ("Masuk", "08.01–08.30", "1,00%"),
            ("Masuk", "08.31–09.00", "1,25%"),
            ("Masuk", "> 09.00", "1,50%"),
            ("Tidak finger masuk", "-", "1,50%"),
            ("Pulang Senin–Kamis", "15.00–15.30", "1,50%"),
            ("Pulang Senin–Kamis", "15.31–16.00", "1,25%"),
            ("Pulang Senin–Kamis", "16.01–16.30", "1,00%"),
            ("Pulang Senin–Kamis", "16.31–16.44", "0,50%"),
            ("Tidak finger pulang", "-", "1,55%"),
            ("Tidak masuk", "Tidak finger masuk dan pulang", "3,00%"),
            ("TL", "Tugas luar", "0,00%"),
            ("WFH / W", "Work From Home; dianggap hadir", "0,00%"),
            ("I", "Izin", "3,00%"),
            ("S", "Rawat inap", "0,00%; ditandai kuning"),
        ]
        for column, value in enumerate(("Jenis", "Rentang/Keterangan", "Potongan"), start=1):
            ExcelExporter._header_style(sheet.cell(3, column, value))
        for row_index, row in enumerate(rows, start=4):
            for column, value in enumerate(row, start=1):
                ExcelExporter._body_style(sheet.cell(row_index, column, value))
        sheet["A20"] = "Perlu keputusan sebelum produksi"
        sheet["A20"].font = Font(bold=True, color="B42318")
        sheet["A21"] = (
            "Rentang 07.31–07.39; pulang sebelum 15.00; rentang Jumat sebelum 12.00; "
            "serta status S tanpa bukti rawat inap."
        )
        sheet.merge_cells("A21:C22")
        sheet["A21"].alignment = Alignment(wrap_text=True, vertical="top")
        sheet.column_dimensions["A"].width = 28
        sheet.column_dimensions["B"].width = 45
        sheet.column_dimensions["C"].width = 25

    @staticmethod
    def _excel_percent(value: Decimal) -> float:
        return float(value / Decimal("100"))

    @staticmethod
    def _header_style(cell) -> None:
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)

    @staticmethod
    def _body_style(cell) -> None:
        cell.font = Font(color=NAVY)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)
