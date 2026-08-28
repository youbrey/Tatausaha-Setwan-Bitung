from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from difflib import SequenceMatcher
from importlib.resources import as_file, files
from pathlib import Path

from openpyxl import load_workbook

from tpp_finger_scan.domain.models import Employee


_WHITESPACE = re.compile(r"\s+")
_NON_NAME = re.compile(r"[^A-Z0-9 ]+")
_TITLE_PREFIX = re.compile(
    r"^(?:DRS|DRA|DR|IR|PROF|HJ|H|MGR|MAG)\s+",
    flags=re.IGNORECASE,
)


def normalize_employee_name(value: str) -> str:
    """Normalisasi nama finger scan dan daftar PNS tanpa gelar akademik."""

    text = unicodedata.normalize("NFKD", value or "")
    text = "".join(character for character in text if not unicodedata.combining(character))
    text = text.upper().strip()
    text = text.split(",", 1)[0]
    text = text.replace(".", " ").replace("-", " ")
    text = _NON_NAME.sub(" ", text)
    text = _WHITESPACE.sub(" ", text).strip()
    while True:
        stripped = _TITLE_PREFIX.sub("", text).strip()
        if stripped == text:
            break
        text = stripped
    return text


@dataclass(frozen=True, slots=True)
class EmployeeReference:
    name: str
    position: str
    nip: str = ""
    rank: str = ""


class EmployeeMaster:
    """Daftar PNS/jabatan bawaan yang dibaca lokal dari workbook referensi."""

    RESOURCE_NAME = "DAFTAR_PNS_DAN_JABATAN.xlsx"

    def __init__(self, workbook_path: str | Path | None = None) -> None:
        self.workbook_path = Path(workbook_path).resolve() if workbook_path else None
        self._records: list[EmployeeReference] | None = None

    @property
    def records(self) -> list[EmployeeReference]:
        if self._records is None:
            self._records = self._load_records()
        return list(self._records)

    def _load_records(self) -> list[EmployeeReference]:
        if self.workbook_path is not None:
            return self._read_workbook(self.workbook_path)

        resource = files("tpp_finger_scan.resources").joinpath(self.RESOURCE_NAME)
        with as_file(resource) as resource_path:
            return self._read_workbook(resource_path)

    @staticmethod
    def _read_workbook(path: str | Path) -> list[EmployeeReference]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook[workbook.sheetnames[0]]
            records: list[EmployeeReference] = []
            pending_index: int | None = None
            for number, name_value, position_value in worksheet.iter_rows(
                min_row=1,
                min_col=1,
                max_col=3,
                values_only=True,
            ):
                if isinstance(number, (int, float)) and name_value and position_value:
                    records.append(
                        EmployeeReference(
                            name=_WHITESPACE.sub(" ", str(name_value)).strip(),
                            position=_WHITESPACE.sub(" ", str(position_value)).strip(),
                        )
                    )
                    pending_index = len(records) - 1
                    continue

                if pending_index is None or not name_value:
                    continue
                detail = _WHITESPACE.sub(" ", str(name_value)).strip()
                current = records[pending_index]
                if detail.upper().startswith("NIP"):
                    nip = re.sub(r"\D", "", detail)
                    records[pending_index] = EmployeeReference(
                        name=current.name,
                        position=current.position,
                        nip=nip,
                        rank=current.rank,
                    )
                elif not current.rank:
                    records[pending_index] = EmployeeReference(
                        name=current.name,
                        position=current.position,
                        nip=current.nip,
                        rank=detail,
                    )
            return records
        finally:
            workbook.close()

    def find(self, employee_name: str) -> EmployeeReference | None:
        target = normalize_employee_name(employee_name)
        if not target:
            return None

        exact = {
            normalize_employee_name(record.name): record
            for record in self.records
        }
        if target in exact:
            return exact[target]

        scored = [
            (
                SequenceMatcher(None, target, normalized).ratio(),
                record,
            )
            for normalized, record in exact.items()
        ]
        if not scored:
            return None
        score, candidate = max(scored, key=lambda item: item[0])
        return candidate if score >= 0.90 else None

    def resolve_positions(self, employees: list[Employee] | tuple[Employee, ...]) -> dict[str, str]:
        positions: dict[str, str] = {}
        for employee in employees:
            match = self.find(employee.name)
            if match and match.position:
                positions[employee.finger_id] = match.position
        return positions
