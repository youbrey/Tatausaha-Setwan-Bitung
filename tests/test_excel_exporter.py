from __future__ import annotations

import tempfile
import unittest
from datetime import date, time
from pathlib import Path

from openpyxl import load_workbook

from tpp_finger_scan.application.services import AttendanceApplicationService, RecapSession
from tpp_finger_scan.domain.models import AttendanceEntry, AttendanceState, Employee, ImportResult
from tpp_finger_scan.domain.rules import DeductionEngine
from tpp_finger_scan.infrastructure.excel_exporter import ExcelExporter


class ExcelExporterTests(unittest.TestCase):
    def test_workbook_contains_summary_and_detail(self) -> None:
        employee = Employee("101", "PEGAWAI UJI")
        entry = AttendanceEntry(
            employee,
            date(2026, 8, 24),
            "08:01-16:45",
            time(8, 1),
            time(16, 45),
            AttendanceState.COMPLETE,
            1,
        )
        result = ImportResult(
            Path("contoh.pdf"),
            "a" * 64,
            date(2026, 8, 24),
            date(2026, 8, 24),
            [employee],
            [entry],
        )
        session = RecapSession(result, [DeductionEngine().calculate(entry)])
        with tempfile.TemporaryDirectory() as directory:
            target = ExcelExporter().export(session, Path(directory) / "rekap.xlsx")
            workbook = load_workbook(target, read_only=True, data_only=False)
            self.assertEqual(workbook.sheetnames, [
                "Rekap Per Pegawai",
                "Master Pegawai",
                "Ringkasan",
                "Detail Harian",
                "Catatan Aturan",
            ])
            self.assertEqual(workbook["Rekap Per Pegawai"]["C10"].value, 0.01)
            self.assertEqual(workbook["Rekap Per Pegawai"]["E10"].value, "=SUM(B10:D10)")
            self.assertEqual(workbook["Ringkasan"]["G8"].value, 0.01)
            self.assertEqual(workbook["Detail Harian"]["N2"].value, "Terlambat")

    def test_employee_recap_keeps_both_deductions_when_checkout_is_missing(self) -> None:
        employee = Employee("120", "YUNIKY RAINTUNG")
        entry = AttendanceEntry(
            employee,
            date(2026, 8, 10),
            "08:31-",
            time(8, 31),
            None,
            AttendanceState.MISSING_OUT,
            1,
        )
        result = ImportResult(
            Path("contoh.pdf"),
            "b" * 64,
            date(2026, 8, 10),
            date(2026, 8, 10),
            [employee],
            [entry],
        )
        calculation = DeductionEngine().calculate(entry)
        session = RecapSession(
            result,
            [calculation],
            employee_positions={"120": "PETUGAS PROTOKOL KOMISIII"},
        )
        with tempfile.TemporaryDirectory() as directory:
            target = ExcelExporter().export(session, Path(directory) / "rekap-yuniky.xlsx")
            workbook = load_workbook(target, read_only=True, data_only=False)
            recap = workbook["Rekap Per Pegawai"]
            self.assertIsNone(recap["B10"].value)
            self.assertEqual(recap["C10"].value, 0.0125)
            self.assertEqual(recap["D10"].value, 0.0155)
            self.assertEqual(recap["E10"].value, "=SUM(B10:D10)")
            self.assertEqual(recap["B3"].value, '=IF(\'Master Pegawai\'!C2="",": -",": "&\'Master Pegawai\'!C2)')


if __name__ == "__main__":
    unittest.main()
